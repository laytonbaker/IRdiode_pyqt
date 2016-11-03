# -*- coding: utf-8 -*-
"""
Created on Thu Jul 31 16:54:13 2014

@author: laytonbaker
"""
from __future__ import division
#import sys
import re
import collections
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
matplotlib.rcParams.update({'font.size': 11})
from openpyxl import load_workbook
from PyQt5.QtWidgets import QMainWindow, QSplitter
from PyQt5.QtCore import pyqtSignal
from guidata.dataset.datatypes import (DataSet, BeginTabGroup, EndTabGroup,
                                       BeginGroup, EndGroup, ObjectItem)
from guidata.dataset.dataitems import (FloatItem, IntItem, BoolItem, ChoiceItem,
                             MultipleChoiceItem, ImageChoiceItem, FilesOpenItem,
                             StringItem, TextItem, ColorItem, FileSaveItem,
                             FileOpenItem, DirectoryItem, FloatArrayItem)
from guidata.dataset.qtwidgets import DataSetShowGroupBox, DataSetEditGroupBox
from guidata.configtools import get_icon
from guidata.qthelpers import create_action, add_actions, get_std_icon

#----------------------------------------------------------------------------
'''
VIEW CLASSES -- Qt GUI interface using guidata and PyQt
'''

class IV_ParamSetA(DataSet):
    ivfiles = FilesOpenItem("IV Data Files", "txt", 
                          basedir=r'C:\Users\laytonba\Documents\JPL_Work\389G\Test_devices',
                          help='Select IV txt file')
    sizefile = FileOpenItem("Size File", ("xlsm", "xls"), 
                          basedir=r'C:\Users\laytonba\Documents\JPL_Work\389G\Test_devices',
                          help='Select pin sizes txt file')    
    savedir = DirectoryItem("Save Directory",                          
                          help='Select picture save directory')
    sizefiletype = ChoiceItem("LCC pins", 
                              [(28, '28-pin setup'), (68, '68-pin setup')])                      
  
    # Plots selection sub-group
    p_group = BeginGroup("Plot Selection")
    plotallbool = BoolItem("plot all devices", default=False)
    plotsizebool = BoolItem("plot IV's for specific size", default=True)
    selectedsize = IntItem("Diode Size (um)", default=250, min=0)
    plotpabool = BoolItem("plot perimeter to area ratio", default=False)
    plotrankbool = BoolItem("plot device rank at Von", default=False)
    bestdevicebool = BoolItem("Find best devices at Von", default=False)
    Von = FloatItem("Turn On (V)", default=-0.1)
    Tscanbool = BoolItem("plot Tscan for listed devices", default=False)
    Tscan_devices = StringItem("Devices", default='1,2,3')
    filterIVbool = BoolItem("filter high/low devices", default=False)
    filterIVlow = FloatItem("low current cutoff", default=1e-8)
    filterIVhigh = FloatItem("high current cutoff", default=1e-2)
    saveplotbool = BoolItem("Save plots to directory", default=False)
    _p_group = EndGroup("Plot Selection")
    
    # Plots formatting subgroup
    t_group = BeginTabGroup("T group")
    a_group = BeginGroup("Plot range")    
    j1 = FloatItem("J min", default=1e-8)
    j2 = FloatItem("J max", default=1)
    v1 = FloatItem("V min", default=-1.5)
    v2 = FloatItem("V max", default=1.5)
    pa1 = FloatItem("PA min", default=10)
    pa2 = FloatItem("PA max", default=500)
    _a_group = EndGroup("Plot Format")
    b_group = BeginGroup("Annotations")
    titlebool = BoolItem("show title")
    titletext = StringItem("Title")
    sizetextbool = BoolItem("indicate device sizes on plot", default=True)
    tempbool = BoolItem("indicate temperature on plot", default=False)
    ptoa_colorbool = BoolItem("ptoa plot in different colors", default=False)
    ranksizebool = BoolItem("rank plot for selected size only", default=False)
    _b_group = EndGroup("Annotations")
    _t_group = EndTabGroup("T group")

class IV_Window(QMainWindow):
#    """Example"""

    def __init__(self):
        #super(IV_Window, self).__init__()
        self.initUI()
        
    def initUI(self):   
        QMainWindow.__init__(self)
        self.setWindowIcon(get_icon('python.png'))
        self.setWindowTitle("28-pin dark current plotting")
        
        # Instantiate dataset-related widgets:
        self.groupbox1 = DataSetEditGroupBox("Dark current density",
                                             IV_ParamSetA,
                                             comment='Files and plot parameters')
        self.groupbox1.SIG_APPLY_BUTTON_CLICKED.connect(self.plotIV)

        # OLD PyQt4 signal for plot function call
        #        self.connect(self.groupbox1, pyqtSignal("apply_button_clicked()"),
        #                     self.plotIV)

        
        splitter = QSplitter(self)
        splitter.addWidget(self.groupbox1)

        self.setCentralWidget(splitter)
        self.setContentsMargins(10, 5, 10, 5)
        
        # File menu
        file_menu = self.menuBar().addMenu("File")
        quit_action = create_action(self, "Quit",
                                    shortcut="Ctrl+Q",
                                    icon=get_std_icon("DialogCloseButton"),
                                    tip="Quit application",
                                    triggered=self.close)
        add_actions(file_menu, (quit_action, ))
    
#----------------------------------------------------------------------------
# MAIN CONTROLLER -- function to instantiate DiodeIV, set attrs', and run
# data processing functions on ivdata

    def setParams(self, instance):
        '''
        text
        '''
        DS1 = self.groupbox1.dataset
        
        instance.v1 = DS1.v1
        instance.v2 = DS1.v2
        instance.j1 = DS1.j1
        instance.j2 = DS1.j2
        instance.pa1 = DS1.pa1
        instance.pa2 = DS1.pa2
        instance.titlebool = DS1.titlebool
        instance.titletext = DS1.titletext
        instance.sizetextbool = True
        instance.filterIVbool = DS1.filterIVbool
        instance.filterIVhigh = DS1.filterIVhigh
        instance.filterIVlow =DS1.filterIVlow
        instance.ptoa_colorbool = DS1.ptoa_colorbool
        
        instance.plotallbool = DS1.plotallbool
        instance.plotsizebool = DS1.plotsizebool
        instance.plotpabool = DS1.plotpabool
        instance.plotrankbool = DS1.plotrankbool
        instance.bestdevicebool = DS1.bestdevicebool
        
        instance.saveplotbool = DS1.saveplotbool
        instance.savedir = DS1.savedir
        
        instance.Tscanbool = DS1.Tscanbool
        instance.Tscan_devices = DS1.Tscan_devices.encode('ascii').split(',')        
        #data and file parameters        
        instance.sizefiletype = DS1.sizefiletype
        instance.ivfiles = DS1.ivfiles
        instance.sizefile = DS1.sizefile
        
        instance.selectedsize = DS1.selectedsize
        instance.Von = DS1.Von
        #instance.devicenum = DS1.devicenum

        # diode data        
        instance.sizedata = []
        instance.ivdata = [] 
 

    def plotIV(self):
        '''
        generate DiodeIV instance, read om data, run analysis
        '''
        #initialize DiodeIV instance and set attributes
        iv = DiodeIV()
        self.setParams(iv)
        DS1 = self.groupbox1.dataset
        
        if iv.sizefiletype == 28:
            iv.sizedata = iv.readsize28(iv.sizefile)
            print('read from 28-pin file')
        elif iv.sizefiletype == 68:
            iv.sizedata = iv.readsize68(iv.sizefile)
            print('read from 68-pin file')
        else:
            print('error, bad file selection')
        
        #print(type(iv.sizedata), type(iv.sizedata[:]) )
        
        iv.ivdata = np.genfromtxt(iv.ivfiles[0], skip_header=0)
        
        print('read ivdata from file')
        
        iv.Ta_list, iv.Tb_list = iv.parseTemp(iv.ivfiles)
        print(iv.Ta_list)
        print(iv.Tb_list)        
        
        if DS1.plotallbool:
            iv.plotIV_all(iv.sizedata, iv.ivdata, iv.Von)
            
        if DS1.plotsizebool:
            iv.plotIV_size(iv.sizedata, iv.ivdata, iv.selectedsize, iv.Von)
            
        if DS1.plotpabool:
            iv.plotIV_ptoa(iv.sizedata, iv.ivdata, iv.Von)
            
        if DS1.plotrankbool:
            if DS1.ranksizebool:
                iv.plotIV_rank(iv.sizedata, iv.ivdata, iv.Von, iv.selectedsize)
            else:
                iv.plotIV_rank(iv.sizedata, iv.ivdata, iv.Von)
        
        if DS1.bestdevicebool:
           iv.bestdevice(iv.sizedata, iv.ivfiles, iv.Von)  
           
        if DS1.Tscanbool:
            devices = DS1.Tscan_devices.encode('ascii').split(',')
            print('generate Tscan for devices:', devices)
            for device in devices:
                iv.plotIV_Tscan(iv.sizedata, iv.ivfiles, int(device))
                
        
#----------------------------------------------------------------------------
'''
MAIN MODEL CLASS
'''

class DiodeIV:
#    """Example"""

    def __init__(self):
        #super(IV_Window, self).__init__()
        self.initUI()
        
    def initUI(self):
        # plot variables        
        self.v1 = -1.5
        self.v2 = 1.5
        self.j1 = 1e-9
        self.j2 = 1
        self.pa1 = 10
        self.pa2 = 300        
        
        self.filterIVbool = False
        self.filterIVhigh = 1e-2
        self.filterIVlow = 1e-7
        
        self.biaslabel = r'$\mathbf{bias\,(V)}$'
        self.jdlabel = r'$\mathbf{J_{dark}\hspace{0.5}}  \mathrm{(A/cm^{2})}$'
        self.palabel = r'$\mathbf{perimeter-to-area \hspace{0.5} ratio,\hspace{0.5}} \mathrm{(cm^{-1})}$'
        self.ranklabel = r'$\mathbf{Device}$'
        
        self.plotallbool = True
        self.plotsizebool = True
        self.plotpabool = True
        self.plotrankbool = True
        self.bestdevicebool = True
        self.titlebool = True
        
        self.saveplotbool = False
        self.savedir = ''
        
        self.Tscanbool = False
        self.Tscan_devices = []
        
        self.sizetextbool = True
        self.titletext = 'Title'
        
        self.ptoa_color = False

        #data and file parameters        
        self.sizefiletype = 68
        self.ivfile = 'test'
        self.sizefile = 'test'
        
        self.selectedsize = 250
        self.Von = -0.1
        self.devicenum = 1
        self.Ta_list = None
        self.Tb_list = None

        # diode data        
        self.sizedata = []
        self.ivdata = []        


       
#------------------------------------------------------------------------------
#Data Analysis class methods
       
    def bestdevice(self, sizedata=None, ivfiles=None, Von=None):
        '''
        rank devices at Von for each size and return a list of indices
        corresponding to the ivdata column for the best device of each size
        '''
        if sizedata is None:
            sizedata = self.sizedata
        if ivfiles is None:
            ivfiles = self.ivfiles
#        if T is None:
#            T = 77  #77K Kevlin evualtion temp
        if Von is None:
            Von = self.Von
        
        size_set = set(sizedata) 
        if None in size_set: 
            size_set.remove(None)
        
        size_list = list(size_set)
        size_list.sort()
        #print(size_list)

        Ta_list, Tb_list = self.parseTemp(ivfiles)
        #print(Ta_list)
        #print(Tb_list)
        
        ivfile = ivfiles[0]
        print('Look for best devices at {0}V in file:\n{1}'.format(Von, ivfile))
#        for i, f in enumerate(ivfiles):
#            matchobj = re.search('_Ta\d\d\d', f)
#            Ta = int(matchobj.group(0)[3:6])
#            #print(Ta)
#            if Ta == T: 
#                ivfile = f
#                print('IV file @ {0}K found: {1}'.format(T, ivfile))
#
#        if not ivfile:
#            print('no IV file match found')
#            return
        
        ivdata = np.genfromtxt(ivfile, skip_header=0)
        ivdata_cols = ivdata.shape[1]
#        print('number of ivdata cols: {0}'.format(ivdata_cols))
        Von_row = ivdata[:,0].tolist().index(Von)
        
        for i, size in enumerate(size_list):
            yset = [a+1 for a, b in enumerate(sizedata)
                        if b == size
                        if a+1 <= ivdata_cols-1 
                        if ivdata[Von_row,a+1] > 0]
            Jset = ivdata[Von_row,yset]/self.pixelarea(size)
            D = dict(zip(yset, Jset))
            print('\nSIZE {0} microns'.format(size))
            print('----------------------------------------')            
            #print(D)
            Ds = collections.OrderedDict(sorted(D.items(),
                                              key= lambda t: t[1] ))
            for device, J in zip(Ds.keys(), Ds.values()):
                print('Device {0}: {1:.2e} A/cm^2'.format(device, J))


    def plotIV_TscanMulti(self, sizedata=None, ivfiles=None, DNs=None):
        for DN in DNs:
            self.plotIV_Tscan(sizedata, ivfiles, DN)
        
    def plotIV_Tscan(self, sizedata=None, ivfiles=None, DN=None):
        '''
        '''
        if sizedata is None:
            sizedata = self.sizedata
        if ivfiles is None:
            ivfiles = self.ivfiles
        if DN is None:
            DN = self.devicenum  #77K Kevlin evualtion temp
        
        size = sizedata[DN-1]
        #print(size)
        
        if not size or size <= 0:
            print('device selected does not have proper size data')
            
        Ta_list, Tb_list = self.parseTemp(ivfiles)
#        print(Ta_list)
#        print(Tb_list)
        
        colormap = plt.cm.jet
        cycle = [colormap(i) for i in np.linspace(0, 0.95, len(Ta_list))]
        
        #initiate figure and axes        
        plt.figure(figsize=(5, 4), dpi=150)
        plt.subplots_adjust(left=0.20, right=0.95,bottom=0.15, 
                            top=0.92, hspace=0.3, wspace=0.3)
        ax = plt.subplot2grid((1,1),(0,0))
        
        for i, ivfile in enumerate(ivfiles):
            ivdata = np.genfromtxt(ivfile, skip_header=0)
            plt.semilogy(ivdata[:,0], 
                         ivdata[:,DN]/self.pixelarea(size), 
                         color=cycle[i], alpha=0.8,
                         label=r'{0}/{1}K'.format(Ta_list[i], Tb_list[i]) )
            
        plt.xlabel(self.biaslabel, fontsize = 18)
        plt.ylabel(self.jdlabel, fontsize=18)
        plt.xlim(self.v1, self.v2)
        plt.ylim(self.j1, self.j2)
        
        plt.grid(b=None, which='major', axis='both', color='gray')
        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontsize(11)
        if self.titlebool == True:               
            plt.title(self.titletext, fontsize = 16)
        if self.sizetextbool:
            plt.text(self.v1+0.05, self.j1*2, 
                     r'device {0}: {1} microns'.format(DN, size),
                     fontsize = 10 ) 
        #plt.text(-1.3, 1e-7, 'IQE2400881 DW2')
        handles, labels = plt.gca().get_legend_handles_labels()
        by_label = collections.OrderedDict(zip(labels, handles))
        plt.legend(by_label.values(), by_label.keys(),
                   fontsize = 10,
                   ncol=1, loc='lower right', 
                   labelspacing=0.0,
                   handletextpad=0.0, handlelength=1.5,
                   fancybox=True, shadow=True)
        if self.saveplotbool:
            picname = self.savedir + '\\{0} d{1}-{2} Tscan.png'.format(self.titletext, DN, size)
#            print(picname)             
            plt.savefig(picname)        
        plt.show()

  
    def plotIV_all(self, sizedata=None, ivdata=None, Von=None):
        '''
        Plot IV's for all device sizes
        '''
        if sizedata is None:
            sizedata = self.sizedata
        if ivdata is None:
            ivdata = self.ivdata        
        if Von is None:
            Von = self.Von
        
        size_set = set(sizedata)
        
        if None in size_set: 
            size_set.remove(None)
        #print(size_set)
        size_list = list(size_set)
        size_list.sort()
        #print(size_list)
        #assert type(ivdata) is np.ndarray 
        Von_row = ivdata[:,0].tolist().index(Von)
         
        #intiiate colormap sequence
        colormap = plt.cm.spectral
        cycle = [colormap(i) for i in np.linspace(0, 0.9, len(size_list))]
        
        #initiate figure and axes        
        plt.figure(figsize=(5, 4), dpi=150)
        plt.subplots_adjust(left=0.20, right=0.95,bottom=0.15, 
                            top=0.92, hspace=0.3, wspace=0.3)
        ax = plt.subplot2grid((1,1),(0,0))
        
        ivdata_cols = ivdata.shape[1]
        for i, size in enumerate(size_list):
            if self.filterIVbool:
                yset = [a+1 for a, b in enumerate(sizedata)
                        if b == size
                        if a+1 <= ivdata_cols-1    
                        if ivdata[Von_row,a+1]/self.pixelarea(size) > self.filterIVlow
                        if ivdata[Von_row,a+1]/self.pixelarea(size) < self.filterIVhigh]       
            else:
                yset = [a+1 for a, b in enumerate(sizedata)
                        if b == size
                        if a+1 <= ivdata_cols-1 ]
            
            #print(yset)
            if len(yset) > 0:                     
                plt.semilogy(ivdata[:,0], 
                             ivdata[:,yset]/self.pixelarea(size), 
                             color=cycle[i], alpha=0.8,
                             label=r'{0}'.format(size) )
        
        plt.xlabel(self.biaslabel, fontsize = 18)
        plt.ylabel(self.jdlabel, fontsize=18)
        plt.xlim(self.v1, self.v2)
        plt.ylim(self.j1, self.j2)
        
        plt.grid(b=None, which='major', axis='both', color='gray')
        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontsize(11)
        if self.titlebool == True:               
            plt.title(self.titletext, fontsize = 18)
        
        if self.Ta_list:
            plt.text(self.v1+0.05, self.j1*2,
                     r'{0}/{1}K'.format(self.Ta_list[0], self.Tb_list[0]) ) 
        
        handles, labels = plt.gca().get_legend_handles_labels()
        by_label = collections.OrderedDict(zip(labels, handles))
        plt.legend(by_label.values(), by_label.keys(),
                   title = 'diode sizes',
                   fontsize = 11,
                   ncol=1, loc='lower right', 
                   labelspacing=0.0,
                   handletextpad=0.0, handlelength=1.5,
                   fancybox=True, shadow=True)
        
        if self.saveplotbool:
            picname = self.savedir + '\\{0} All.png'.format(self.titletext)
            #print(picname)             
            plt.savefig(picname)
        plt.show()

    def plotIV_size(self, sizedata=None, ivdata=None, size=None, Von=None):
        '''
        Plot IV's for selected diode size
        opional ARG: diode size in microns to be plotted
        '''  
        if sizedata is None:
            sizedata = self.sizedata
        if ivdata is None:
            ivdata = self.ivdata
        if size is None:
            size = self.selectedsize
        if Von is None:
            Von = self.Von    
        
        Von_row = ivdata[:,0].tolist().index(Von)
        
        #assert type(ivdata) is np.ndarray 
        ivdata_cols = ivdata.shape[1]  
        if self.filterIVbool:
            yset = [a+1 for a, b in enumerate(sizedata)
                    if b == size
                    if a+1 <= ivdata_cols-1
                    if ivdata[Von_row,a+1]/self.pixelarea(size) > self.filterIVlow
                    if ivdata[Von_row,a+1]/self.pixelarea(size) < self.filterIVhigh]       
        else:
            yset = [a+1 for a, b in enumerate(sizedata)
                    if b == size
                    if a+1 <= ivdata_cols-1 ]
         
        #Plot IV's for specific device size
        plt.figure(figsize=(5, 4), dpi=150)
        plt.subplots_adjust(left=0.20, right=0.95,bottom=0.15, 
                            top=0.92, hspace=0.3, wspace=0.3)
        ax = plt.subplot2grid((1,1),(0,0))
    
        #print(yset)
        if len(yset) > 0:          
            plt.semilogy(ivdata[:,0], 
                         ivdata[:,yset]/self.pixelarea(size), 
                         'r', alpha=0.8)
    
        plt.xlabel(self.biaslabel, fontsize = 18)
        plt.ylabel(self.jdlabel, fontsize=18)
        plt.xlim(self.v1, self.v2)
        plt.ylim(self.j1, self.j2)
        
        plt.grid(b=None, which='major', axis='both', color='gray')
        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontsize(11)
        if self.titlebool == True:               
            plt.title(self.titletext, fontsize = 18)
        if self.sizetextbool == True:
            plt.text(self.v1+0.05, self.j1*2, 
                     r'{0} micron devices, {1}/{2}K'.format(size, self.Ta_list[0], self.Tb_list[0]) ) 
        #plt.text(-1.3, 1e-7, 'IQE2400881 DW2')
        if self.saveplotbool:
            picname = self.savedir + '\\{0} {1}microns.png'.format(self.titletext, size)
            #print(picname)             
            plt.savefig(picname)        
        
        plt.show()

    def plotIV_ptoa(self, sizedata=None, ivdata=None, Von=None):
        '''
        plot dark currenty density @ Von versus perimeter to are ratio
        '''
        if sizedata is None:
            sizedata = self.sizedata
        if ivdata is None:
            ivdata = self.ivdata
        if Von is None:
            Von = self.Von
        
        Von_row = ivdata[:,0].tolist().index(Von)
        size_set = set(sizedata)
        if None in size_set: 
            size_set.remove(None)
        #print(size_set)
        size_list = list(size_set)
        size_list.sort()
        #Plot J versus perimeter to area ratio
        plt.figure(figsize=(5, 4), dpi=150)
        plt.subplots_adjust(left=0.20, right=0.95, bottom=0.15, 
                            top=0.92, hspace=0.3, wspace=0.3)
        ax = plt.subplot2grid((1,1),(0,0))
        
        #assert type(ivdata) is np.ndarray 
        ivdata_cols = ivdata.shape[1]  
        if self.ptoa_colorbool:
            colormap = plt.cm.spectral
            cycle = [colormap(i) for i in np.linspace(0, 0.9, len(size_list))]
            for i, size in enumerate(size_list):
                yset = [a+1 for a, b in enumerate(sizedata) 
                        if b == size
                        if a+1 <= ivdata_cols-1 ]
                for y in yset:                       
                    plt.semilogy(self.ptoa(size), 
                                 ivdata[Von_row, y]/self.pixelarea(size), 
                                 color=cycle[i], marker='o', alpha=0.8,
                                 label=r'{0}'.format(size) )

        else:          
            for a, size in enumerate(sizedata):            
                if size > 0 and a+1 <= ivdata_cols-1:
                    plt.semilogy(self.ptoa(size), 
                                 ivdata[Von_row,a+1]/self.pixelarea(size),
                                 marker='o', color='red', alpha=0.75)
        
        plt.xlabel(self.palabel, fontsize=14)
        plt.ylabel(self.jdlabel, fontsize=18)
        plt.xlim(self.pa1, self.pa2)
        plt.ylim(self.j1, self.j2)
        
        plt.grid(b=None, which='major', axis='both', color='gray')
        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontsize(11)

        if self.titlebool == True:               
            plt.title(self.titletext, fontsize = 16)

        if self.sizetextbool == True:
            plt.text(self.pa1+5, self.j1*2, 
                     r'{0}/{1}K'.format(self.Ta_list[0], self.Tb_list[0]) ) 
        
        plt.text(20, self.j2*0.3, r'Turn On Bias = {} V'.format(Von) )
        if self.ptoa_colorbool:
            handles, labels = plt.gca().get_legend_handles_labels()
            by_label = collections.OrderedDict(zip(labels, handles))
            plt.legend(by_label.values(), by_label.keys(), 
                       title = 'diode sizes',
                       fontsize = 11,
                       ncol=1, loc='lower right', 
                       labelspacing=0.0, markerscale = 0.8,
                       handletextpad=0.0, handlelength=1.5,
                       fancybox=True, shadow=True)

        if self.saveplotbool:
            picname = self.savedir + '\\{0} ptoa.png'.format(self.titletext)
            #print(picname)             
            plt.savefig(picname)        
        plt.show()

    def plotIV_rank(self, sizedata=None, ivdata=None, Von=None,
                    selectedsize=None):
        '''
        plot dark currenty density rank plot for all devices
        ARGS: if diode size in microns is supplied, rank plot is performed only
        for diode of 'size' size
        '''
        if sizedata is None:
            sizedata = self.sizedata
        if ivdata is None:
            ivdata = self.ivdata
        if Von is None:
            Von = self.Von
        print('gen rank plot at Von: {0}'.format(Von))            
        Von_row = ivdata[:,0].tolist().index(Von)

        #assert type(ivdata) is np.ndarray 
        ivdata_cols = ivdata.shape[1]
        if selectedsize is None:
            yset, size_set = zip(*[(a+1,b) for a,b in enumerate(self.sizedata)
                                if b > 0
                                if a+1 <= ivdata_cols-1 ])
        else:
            yset, size_set = zip(*[(a+1,b) for a,b in enumerate(self.sizedata)  
                                if b == selectedsize
                                if a+1 <= ivdata_cols-1 ])
        #print(yset)
        #print(size_set)        
        JVset = [ivdata[Von_row,y]/self.pixelarea(size) 
                  for y, size in zip(yset, size_set)
                  if ivdata[Von_row,y] > 0]
        JVset.sort(reverse=False)              
        #print(JVset)
        #print(len(yset), len(size_set), len(JVset))
        #Plot J versus perimeter to area ratio
        plt.figure(figsize=(5, 4), dpi=150)
        plt.subplots_adjust(left=0.20, right=0.95, bottom=0.15, 
                            top=0.92, hspace=0.3, wspace=0.3)
        ax = plt.subplot2grid((1,1),(0,0))
       
        for i, J in enumerate(JVset):
            plt.semilogy(i+1, J,
                         marker='o', color='red', alpha=0.75)
        
        plt.xlabel(self.ranklabel, fontsize = 16)
        plt.ylabel(self.jdlabel, fontsize= 18)
        plt.xlim(0, len(JVset)+1)
        plt.ylim(self.j1, self.j2)
        
        plt.grid(b=None, which='major', axis='both', color='gray')
        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontsize(11)
        if self.titlebool == True:               
            plt.title(self.titletext, fontsize = 16)
        plt.text(0.5, self.j2*0.3, r'Turn On Bias = {} V'.format(Von) )
        plt.text(0.5, self.j1*2, 
                 r'{0}/{1}K'.format(self.Ta_list[0], self.Tb_list[0]) ) 
        #plt.text(-1.3, 1e-7, 'IQE2400881 DW2')

        if self.saveplotbool:
            picname = self.savedir + '\\{0} rank.png'.format(self.titletext)
            print(picname)             
            plt.savefig(picname)
        plt.show()

    def plot_ArrMulti(self, sizedata=None, ivfiles=None, Von=None, DNs=None):
        for DN in DNs:
            self.plot_Arr(sizedata, ivfiles, Von, DN)        
    
    def plot_Arr(self, sizedata=None, ivfiles=None, Von=None, DN=None):
        '''
        plot ln(Jd/T^3) vs. 1000/T
        '''
        if sizedata is None:
            sizedata = self.sizedata
        if ivfiles is None:
            ivfiles = self.ivfiles
        if DN is None:
            DN = self.devicenum  #77K Kevlin evualtion temp
        if Von is None:
            Von = self.Von
            
        size = sizedata[DN-1]

        if not size or size <= 0:
            print('device selected does not have proper size data')
            return
        
        Ta_list, Tb_list = self.parseTemp(ivfiles)
        Ta_narray = np.array(Ta_list, dtype=np.float)
#        print(Ta_list)
#        print(Tb_list)   
        
        #compose list of Jd(Von) for each temperature in Ta_list
        Jd = []
        for i, ivfile in enumerate(ivfiles):
            ivdata = np.genfromtxt(ivfile, skip_header=0)
            Von_row = ivdata[:,0].tolist().index(Von)
            Jd.append(ivdata[Von_row, DN]/self.pixelarea(size))
            
        #initiate figure and axes        
        plt.figure(figsize=(5, 4), dpi=150)
        plt.subplots_adjust(left=0.20, right=0.95,bottom=0.15, 
                            top=0.92, hspace=0.3, wspace=0.3)
        ax = plt.subplot2grid((1,1),(0,0))
        
        plt.scatter(1000.0/Ta_narray, Jd)
        plt.yscale('log')
        plt.ylim(1e-6, 1)
            
        plt.xlabel(r'$\mathbf{1000/T\hspace{0.5}}  \mathrm{(K^{-1})}$', fontsize = 18)
        plt.ylabel(r'$\mathbf{ln(J_{d}/T^{3})\hspace{0.5}}$', fontsize=18)
        
        plt.grid(b=None, which='major', axis='both', color='gray')
        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontsize(11)
        if self.titlebool == True:               
            plt.title(self.titletext, fontsize = 16)
        if self.sizetextbool:
            plt.text(self.v1+0.05, self.j1*2, 
                     r'device {0}: {1} microns'.format(DN, size),
                     fontsize = 10 ) 
        #plt.text(-1.3, 1e-7, 'IQE2400881 DW2')
        if self.saveplotbool:
            picname = self.savedir + '\\{0} d{1}-{2} Arr.png'.format(self.titletext, DN, size)
#            print(picname)             
            plt.savefig(picname)        
        plt.show()    

    def saveJV(self, filename=None):
        '''
        print dark current density to file
        '''
        pass
        


    def pixelarea(self, pitch):
        '''
        Calculates pixel area from pixel pitch
        ARGS: pixel pitch in microns (um)
        RETURNS: pixel area in centimeters (cm)
        '''
        return (pitch**2)/1.0e8
    
    def ptoa(self, pitch):
        '''
        Calculates pixel perimeter to area ratio in cm^-1
        ARGS: pixel pitch in microns (um)
        RETURNS: pixel perimeter to area ratio (cm^-1)
        '''    
        return 40000.0/pitch
        
    def parseIV(self, sizedata, ivdata, size, Von):
        '''
        finds indices of proper size data columns in ivdata
        finds row index of turn-on voltage
        '''
        xlist = [a for a, b in enumerate(sizedata) if b == size]       
        ylist = [x+1 for x in xlist]
        q = ivdata[:,0].tolist().index(Von)
        return xlist, ylist, q

    def readIV(self, filename):
        self.ivfile = filename        
        self.ivdata = np.genfromtxt(filename, skip_header=0)
    
    def readsize28(self, filename):
        #from openpyxl import load_workbook
        self.sizefile = filename
        wb = load_workbook(filename)
        ws = wb['Sheet1']
        #self.sizedata = []
        sizes = []
        for row in ws.range('B2:B21'):
            for cell in row:
                sizes.append(cell.value)
        #self.sizedata = sizes
        return sizes
        
    def readsize68(self, filename):
        #from openpyxl import load_workbook
        self.sizefile = filename
        wb = load_workbook(filename)
        ws = wb['pinsizes']
        #self.sizedata = []
        sizes = []
        flip_list = [False, True, True, True, True, False, False, False]    
        cells_list = ['Q31:W31', 'Z21:Z28', 'Z12:Z18', 'P9:W9',
                      'G9:M9', 'D12:D19', 'D22:D28', 'G31:N31']
        for flip, cells in zip(flip_list, cells_list):    
            temp = []         
            for row in ws.range(cells):
                for cell in row:
                    temp.append(cell.value)
            if flip: 
                temp.reverse()
            #print(temp)
            sizes.extend(temp)
        #self.sizedata = sizes
        return sizes    

    def parseTemp(self, filenames):
        '''
        pure function to extract test temperature from filenames
        receives a list of filenames and returns two lists of temperatures
        '''
        Ta_list = []
        Tb_list = []
        for filename in filenames:
#            print('filename is: ', filename)
#            print(type(filename))
            matchobj = re.search('_Ta\d\d\d_Tb\d\d\d', filename.encode('ascii'))
#            print(matchobj.group(0), type(matchobj.group(0)))
            Ta = int(matchobj.group(0)[3:6])
            Tb = int(matchobj.group(0)[9:12])    
            Ta_list.append(Ta)
            Tb_list.append(Tb)
        return Ta_list, Tb_list


if __name__ == '__main__':

    runreal = True
    
    if runreal:
        from PyQt5.QtWidgets import QApplication
        import sys
        try:
            app = QApplication(sys.argv)
        except RuntimeError:
            app = QApplication.instance()
        iv = IV_Window()
        iv.show()
        sys.exit(app.exec_())
    
# -------------------------------------------
# ELSE if for dirty testing...
    else:
        import glob
        #filedir = 'C:\\RawData\\20140729_IET111481_IV_Tscan\\'
        filedir = 'C:\\RawData\\20140804_IET111498_A3EB01\\'
        ivfiles = glob.glob(filedir + '*_I.txt')
        #sizefile = glob.glob(filedir + '*.xlsm')
        #sizefile= 'C:/RawData/20140729_IET111481_IV_Tscan/IET111481 A4IA02 size.xlsm'
        sizefile= 'C:/RawData/20140804_IET111498_A3EB01/IET111498 A3EB01 sizes.xlsm'
        #print(sizefile)
            
        iv = DiodeIV()
        sizes = iv.readsize68(sizefile)
        ivdata = np.genfromtxt(ivfiles[0], skip_header=0)
        iv.sizedata = sizes
        iv.ivdata = ivdata
        iv.Ta_list, iv.Tb_list = iv.parseTemp(ivfiles)
        
        iv.saveplotbool = True
        iv.titletext = 'IQE111498'
        iv.savedir = filedir
#        print(iv.__class__.__name__)
#        for key in iv.__dict__.keys():
#            print(key, '=>', iv.__dict__[key])
       

#        iv.plotIV_all()
#        iv.plotIV_size(size=250)
#        iv.plotIV_ptoa()
#        iv.plotIV_rank(Von=0.4)
#        iv.bestdevice(sizes, ivfiles, -0.1)
#        iv.plotIV_Tscan(sizes, ivfiles, 6)
#        iv.plotIV_TscanMulti(sizes, ivfiles, [6, 18, 3]) 
        iv.plot_Arr(sizes, ivfiles, 0.4, 6)